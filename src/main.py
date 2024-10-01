import asyncio
import operator
import os
import re

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.async_api import async_playwright

from enums import ACESSO_CIDADAO_ADMIN, GRUPO_E_SERVIDOR
from helpers import retornar_linhas_tabela

load_dotenv()


async def main():
    sufixo_permissao = " - PCIES"
    excel_nome = "permissao-edocs.xlsx"
    excel_coluna_minima = 1
    excel_coluna_maxima = 5
    excel_linha_minima = 2
    excel_coluna_nome = 0
    excel_coluna_cpf = 1
    excel_coluna_lotacao = 2
    excel_coluna_funcao = 3
    excel_coluna_acao = 4

    permissoes_basicas = [
        # "Acessar Caixa de Documentos de Órgão/Setor (Encaminhamentos)"
        "perfil-117",
        # "Acessar Caixas de Processos de Órgão/Setor (Despachar Processo)"
        "perfil-110",
        # "Acessar Documentos Credenciados de Órgão/Setor"
        "perfil-66",
        # "Autuar Processo"
        "perfil-95",
        # "Reabrir Processos Encerrados do Setor"
        "perfil-122"
    ]

    def configurar_proxy():
        proxy_url = os.getenv("PROXY_URL")
        proxy_porta = os.getenv("PROXY_PORTA")
        proxy_usuario = os.getenv("PROXY_USUARIO")
        proxy_senha = os.getenv("PROXY_SENHA")
        proxy_config = None

        if (proxy_url != '' and proxy_porta != '' and
                proxy_usuario != '' and proxy_senha != ''):
            proxy_config = {
                "server": f"{proxy_url}:{proxy_porta}",
                "username": proxy_usuario,
                "password": proxy_senha
            }

        return proxy_config

    def ler_excel():
        diretorio_excel = os.path.join(
            os.getcwd(), "src", "dados", excel_nome)
        wb = load_workbook(diretorio_excel, data_only=True)
        return wb.active

    async def configurar_navegador():
        navegador = await p.chromium.launch(
            headless=False, slow_mo=50, args=['--start-maximized'], proxy=configurar_proxy())
        contexto = await navegador.new_context(no_viewport=True)
        pagina = await contexto.new_page()
        return pagina

    async def buscar_url_acesso_cidadao_admin(pagina):
        botao_acesso_cidadao_admin = pagina.get_by_title(
            "Acessar: Acesso Cidadão (")
        javascript_acesso_cidadao_admin = await botao_acesso_cidadao_admin.get_attribute(
            "onclick")
        url = re.search(
            r"(?P<url>https?://[^\s]+)", javascript_acesso_cidadao_admin).group("url")
        return url

    async def acessar_pagina(pagina, url):
        return await pagina.goto(url)

    async def acessar_pagina_sistemas(pagina, acesso_cidadao_admin):
        await pagina.locator(f"#{acesso_cidadao_admin}").click()
        await pagina.get_by_role("link", name="E-Docs sigades").click()
        await pagina.get_by_role("link", name="Verificar").click()

    async def acessar_pagina_grupos_e_servidores(pagina, acesso_cidadao_admin, grupo_e_servidor):
        await pagina.locator(f"#{acesso_cidadao_admin}").click()
        await pagina.get_by_role("link", name=grupo_e_servidor, exact=True).click()

    async def pesquisar_por_cpf_pagina_sistemas(pagina, cpf):
        await pagina.get_by_role("link", name="Adicionar").click()
        await pagina.get_by_label("CPF ou e-mail").fill(cpf)
        await pagina.get_by_role("button", name="Pesquisar").click()

    async def realizar_login(pagina):
        await pagina.get_by_role("button", name="Login via Acesso Cidadão").click()
        await pagina.get_by_placeholder("CPF").fill(os.getenv("ACESSO_CIDADAO_LOGIN"))
        await pagina.get_by_placeholder("Senha").fill(
            os.getenv("ACESSO_CIDADAO_SENHA"))
        await pagina.get_by_role("button", name="Entrar", exact=True).click()

    async def pesquisar_por_cpf_pagina_servidor(pagina, cpf):
        await pagina.get_by_placeholder("CPF ou e-mail").fill(cpf)
        await pagina.get_by_text("search").click()

    async def acessar_papeis(pagina):
        await pagina.get_by_role("link", name="assignment_ind").click()

    async def verificar_papel(pagina, papel):
        papel_existe = False
        linhas_papeis = await retornar_linhas_tabela(pagina)

        for i in range(await linhas_papeis.count()):
            colunas_papeis = linhas_papeis.nth(i).locator("td")
            for j in range(await colunas_papeis.count()):
                if operator.contains(await colunas_papeis.nth(j).inner_text(), papel):
                    papel_existe = True
                    break

        return papel_existe

    async def cadastrar_papel(pagina, papel):
        await pagina.get_by_role("link", name="Adicionar", exact=True).click()
        await pagina.get_by_text("Selecione um período 1 mês").click()
        await pagina.get_by_role("option", name="ano (máximo)").click()
        await pagina.get_by_label("Nome").fill(papel)
        await pagina.get_by_role("button", name="Adicionar").click()

    async def remover_papel(pagina, papel):
        linhas_papeis = await retornar_linhas_tabela(pagina)
        botao_excluir_papel = None
        posicao_coluna_botoes_acoes = 4

        for i in range(await linhas_papeis.count()):
            colunas_papeis = linhas_papeis.nth(i).locator("td")
            for j in range(await colunas_papeis.count()):
                if operator.contains(await colunas_papeis.nth(j).inner_text(), papel):
                    coluna_botoes_acoes = colunas_papeis.nth(
                        posicao_coluna_botoes_acoes)
                    break

        if coluna_botoes_acoes is not None:
            formulario_acoes = coluna_botoes_acoes.locator("form")
            botao_excluir_papel = formulario_acoes.locator("button")

            if botao_excluir_papel is not None:
                await botao_excluir_papel.click()

    async def cadastrar_lotacao(pagina, funcao, lotacao):
        # https://stackoverflow.com/questions/77014199/python-playwright-want-to-print-all-text-in-table-but-finds-0-rows
        tabela_papeis = pagina.locator("table")
        linhas_papeis = tabela_papeis.locator("tbody tr")

        posicao_coluna_botoes_acoes = 4
        coluna_botoes_acoes = None
        link_editar_papel = None

        for i in range(await linhas_papeis.count()):
            colunas_papeis = linhas_papeis.nth(i).locator("td")
            for j in range(await colunas_papeis.count()):
                if operator.contains(await colunas_papeis.nth(j).inner_text(), funcao):
                    coluna_botoes_acoes = colunas_papeis.nth(
                        posicao_coluna_botoes_acoes)
                    break

        if coluna_botoes_acoes is not None:
            link_editar_papel = coluna_botoes_acoes.locator("a")

            if link_editar_papel is not None:
                await link_editar_papel.click()
                await pagina.get_by_role("link", name="add_circle").click()
                await pagina.get_by_label("Só permitidos").check()
                await pagina.get_by_title(lotacao).locator("#Selecionado").click()
                await pagina.get_by_role("button", name="Selecionar").click()
                await pagina.get_by_label("Voltar").click()

    async def cadastrar_permissoes(pagina, funcao, lotacao):
        linhas_papeis = await retornar_linhas_tabela(pagina)
        linha_papel = None
        papel_lotacao = f"{funcao} ({lotacao})"

        for i in range(await linhas_papeis.count()):
            colunas_papeis = linhas_papeis.nth(i).locator("td")
            for j in range(await colunas_papeis.count()):
                texto = await colunas_papeis.nth(j).inner_text()
                if operator.contains(texto, papel_lotacao):
                    linha_papel = linhas_papeis.nth(i)
                    break

        if linha_papel is not None:
            link_permissoes = linha_papel.locator("a")
            await link_permissoes.click()

            for permissao in permissoes_basicas:
                await pagina.locator(f"#{permissao}").check()

            await pagina.get_by_role("button", name="Salvar").click()

            links = pagina.get_by_role("link", name="Selecionar Orgão")

            for i in range(await links.count()):
                await links.nth(i).click()
                await pagina.get_by_role("textbox").fill(lotacao)
                await pagina.get_by_title(lotacao).locator(
                    "[id=\"node\\.guid\"]").check()
                await pagina.get_by_role("button", name="Salvar").click()
                await pagina.get_by_label("Voltar").click()

            await pagina.get_by_label("Voltar").click()

    async with async_playwright() as p:
        excel = ler_excel()

        pagina = await configurar_navegador()
        await acessar_pagina(pagina, os.getenv("ACESSO_CIDADAO_URL"))
        await realizar_login(pagina)
        url_acesso_cidadao_admin = await buscar_url_acesso_cidadao_admin(pagina)
        await acessar_pagina(pagina, url_acesso_cidadao_admin)

        for linha in excel.iter_rows(values_only=True, min_col=excel_coluna_minima,
                                     max_col=excel_coluna_maxima, min_row=excel_linha_minima):
            nome = linha[excel_coluna_nome]
            cpf = linha[excel_coluna_cpf]
            lotacao = linha[excel_coluna_lotacao]
            lotacao_sigla = lotacao[0:lotacao.index(
                "-")].strip() if lotacao is not None else None
            funcao = linha[excel_coluna_funcao]
            acao = linha[excel_coluna_acao]

            if nome == None or cpf == None or lotacao_sigla == None or funcao == None or acao == None:
                continue

            # await acessar_pagina_grupos_e_servidores(pagina, ACESSO_CIDADAO_ADMIN.GRUPOS_E_SERVIDORES.value, GRUPO_E_SERVIDOR.SERVIDOR.value)
            # await pesquisar_por_cpf_pagina_servidor(pagina, cpf)
            # await acessar_papeis(pagina)

            await acessar_pagina_sistemas(pagina, ACESSO_CIDADAO_ADMIN.SISTEMAS.value)
            await pesquisar_por_cpf_pagina_sistemas(pagina, cpf)
            await cadastrar_permissoes(pagina, funcao, lotacao_sigla)

            # papel_existe = await verificar_papel(pagina, funcao)
            # print("papel_existe ", papel_existe)
            # if not papel_existe:
            #     await cadastrar_papel(pagina, funcao)
            #     await cadastrar_lotacao(pagina, funcao, lotacao_sigla)
            #     await acessar_pagina(pagina, url_acesso_cidadao_admin)
            #     await acessar_pagina_sistemas(pagina, ACESSO_CIDADAO_ADMIN.SISTEMAS.value)
            #     await pesquisar_por_cpf_pagina_sistemas(pagina, cpf)

            #     papel_existe = await verificar_papel(pagina, funcao)
            #     print("papel_existe ", papel_existe)

            #     await cadastrar_permissoes(pagina, f"{funcao} ({lotacao_sigla})", None)
            # else:
            #     await remover_papel(pagina, funcao)

        await pagina.pause()
    # browser.close()

if __name__ == "__main__":
    asyncio.run(main())
